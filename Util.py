import openpyxl
from Bio import SeqIO
import glob
import os

class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_dat = ".dat"
        self.ext_xlsx = ".xlsx"

    """
    reads the fastq files into a dictionary with the file names as keys
    :param
        FASTQ file
        @M00265:331:000000000-CJRDH:1:1101:19061:2012 1:N:0:200
        TGTTTCACAAACAAGACAGGGAACTGGCAGGCACCGAGGCATCTCTGCACCGAGGTGAAACAAGCTGCCATTTCATTACAGGCAAAGCTGAGCAAAAGTAGATATTACAAGACCAGCATGTACTCACCTCTCATGAAGCACTGTGGGTACGAAGGAAATGACTCAAATATGCTGTCTGAAGCCATCGCTTCCTCCTGAAAATGCACCCTCTTCTGAAGGCGGGGGACTCAATGATTTCTTTTACCTTCGGAGCGAAAACCAAGACAGGTCACTGTTTC
        +
        ABBBBFFFFFFAGCFFGFGGGCFFHHGHFGGFHHGGAECGDGHHHHGHFHHGCEGGHGHEHHHFHHHHHHHHHHBGHFHFFHGGEGHHHHEHHHHGHHHHHGHHHHFHHFHGHHHHHGHHHFHHHHHHHHHHHHHEHHHHHHHHHHGGHHGEAEGGHGHHGHHHHFFHHHFHHHHHHHEFGHHHHHGGHHGHHHHHHBFCBGHHHHHGHHGGGGGEB0FEGGAF;DDFFFFFFFFFFFFFFFFFFFFFFF?BFFFFFFFFFFFFFEF?FFFFFFFFFF
        @M00265:331:000000000-CJRDH:1:1101:18745:2023 1:N:0:200
        TCGGTCACAAACAAGACAGGGAACTGGCAGGCACCGAGGCATCTCTGCACCGAGGTGAAACAAGCTGCCATTTCATTACAGGCAAAGCTGAGCAAAAGTAGATATTACAAGACCAGCATGTACTCACCTCTCATGAAGCACTGTGGGTACGAAGGAAATGACTCAAATATGCTGTCTGAAGGCATCGCTTCCTCCTGAAAATGCACCCTCTTCTGAAGGCGGGGGACTCAATGATTTCTTTTACCTTCGGAGCGAAAACCAAGACAGGTCACTGTTTC
        +
        BABBBBBFFFFBGCBGFFGFGCFFHGBGFGGEHFGE?2EGCGGHHHFHFGHGGEGGGGHFHHEFGHHHH3FHHHFHHGG3BFFEAFGFGG3BFBBCGFHHHEGGGHGHFHBFHGHAG0GEGDGHHFHHHHHHBG121DGGHHCHFBAEFHDF/AA0FCGGDDGHCCGHHGFGHFHHHDDFDGCCGHADHHGHHHHHB:0:CCCGFGHGGGGGGGB0B09:A----BBFFFFFFFFFFFFFFFFFFFFFFDDBBBB=BBFFFE??FFFEFFFFFFFBFF
        @M00265:331:000000000-CJRDH:1:1101:19677:2030 1:N:0:200
        GACTTCACAAACAAGACAGGGAACTGGCAGGCACCGAGGCATCTCTGCACCGAGGTGAAACAAGCTGCCATTTCATTACAGGCAAAGCTGAGCAAAAGTAGATATTACAAGACCAGCATGTACTCACCTCTCATGAAGCACTGTGGGTACGAAGGAAATGACTCAAATATGCTGTCTGAAGCCATCGCTTCCTCCTGAAAATGCACCCTCTTCTGAAGGCGGGGGACTCAATGATTTCTTTTACCTTCGGAGCGAAAACCAAGACAGGTCACTGTTTC
        +

    :return
        {   'file name as key' : [FASTQ reads list from the file as values]
            , 'D:/000_WORK/KimHuiKwon/20200327/Fig 4d\\Fig4_d_RUNX1_6GC_rep3.fastq': 
    	        ['TGTTTCACAAACAAGACAGGGAACTGGCAGGCACCGAGGCATCTCTGCACCGAGGTGAAACAAGCTGCCATTTCATTACAGGCAAAGCTGAGCAAAAGTAGATATTACAAGACCAGCATGTACTCACCTCTCATGAAGCACTGTGGGTACGAAGGAAATGACTCAAATATGCTGTCTGAAGCCATCGCTTCCTCCTGAAAATGCACCCTCTTCTGAAGGCGGGGGACTCAATGATTTCTTTTACCTTCGGAGCGAAAACCAAGACAGGTCACTGTTTC'
    	        , 'TCGGTCACAAACAAGACAGGGAACTGGCAGGCACCGAGGCATCTCTGCACCGAGGTGAAACAAGCTGCCATTTCATTACAGGCAAAGCTGAGCAAAAGTAGATATTACAAGACCAGCATGTACTCACCTCTCATGAAGCACTGTGGGTACGAAGGAAATGACTCAAATATGCTGTCTGAAGGCATCGCTTCCTCCTGAAAATGCACCCTCTTCTGAAGGCGGGGGACTCAATGATTTCTTTTACCTTCGGAGCGAAAACCAAGACAGGTCACTGTTTC'
    	        , 'GACTTCACAAACAAGACAGGGAACTGGCAGGCACCGAGGCATCTCTGCACCGAGGTGAAACAAGCTGCCATTTCATTACAGGCAAAGCTGAGCAAAAGTAGATATTACAAGACCAGCATGTACTCACCTCTCATGAAGCACTGTGGGTACGAAGGAAATGACTCAAATATGCTGTCTGAAGCCATCGCTTCCTCCTGAAAATGCACCCTCTTCTGAAGGCGGGGGACTCAATGATTTCTTTTACCTTCGGAGCGAAAACCAAGACAGGTCACTGTTTC'
    	        , ... ]
    	    , 'D:/000_WORK/KimHuiKwon/20200327/Fig 4d\\Fig4_d_RUNX1_6GC_rep3.fastq':
    	        [...]
    	    ...}
    """
    def get_FASTQ_seq_dict(self, sources):
        fastq_dict = {}
        for i in range(len(sources)):
            temp = list(SeqIO.parse(sources[i], "fastq"))
            fastq_dict[sources[i]] = [str(temp[k].seq) for k in range(len(temp))]

        return fastq_dict

    def get_FASTQ_seq_list(self, fastq_path):
        temp = list(SeqIO.parse(fastq_path, "fastq"))
        return [str(temp[k].seq) for k in range(len(temp))]

    def get_FASTQ_seq_with_targt_seq_dict(self, sources, trgt_seq):
        fastq_dict = {}
        for i in range(len(sources)):
            temp = list(SeqIO.parse(sources[i], "fastq"))
            fastq_dict[sources[i]] = [str(temp[k].seq) for k in range(len(temp)) if trgt_seq.upper() in str(temp[k].seq).upper()]

        return fastq_dict

    """
    :return
        result_list = [
                        [INDEX, TTTT_Barcode, Original sequence, Edited sequence]
                        , [CFTR_Forward TTT deletion, TTTTCTGCACGCATACTCTCAT, AAATATCATCTTTGGTGTTTCCTATGATGA, AAATATCATCGGTGTTTCCTATGATGAGTA]
                        ...
                    ]
    """
    def read_tb_txt(self, path):
        result_list = []
        with open(path, "r") as f:
            f.readline().replace("\n", "")
            while True:
                tmp_line = f.readline().replace("\n", "")
                if tmp_line == "":
                    break

                tmp_arr = tmp_line.split("\t")
                result_list.append(tmp_arr)

        return result_list

    """
        :return
            data_list = [
                            [INDEX, TTTT_Barcode, Original sequence, Edited sequence]
                            , [CFTR_Forward TTT deletion, TTTTCTGCACGCATACTCTCAT, AAATATCATCTTTGGTGTTTCCTATGATGA, ""]
                            , [CFTR_Forward TTT deletion, TTTTCTGCACGCATACTCTCAT, "", AAATATCATCGGTGTTTCCTATGATGAGTA]
                            , [CFTR_Forward TTT deletion, TTTTCTGCACGCATACTCTCAT, "", AAATATCATCGGTGTTTCCTATGATGAGTA]
                            ...
                        ]
        """
    def make_list_to_excel(self, result_path, data_list):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="index")
        sheet.cell(row=row, column=2, value='TTTT_Barcode')
        sheet.cell(row=row, column=3, value='Original sequence')
        sheet.cell(row=row, column=4, value='Edited sequence')
        sheet.cell(row=row, column=5, value='FASTQ sequence')

        for val_arr in data_list:
            row += 1
            col = 0
            for val in val_arr:
                col += 1
                sheet.cell(row=row, column=col, value=val)

        workbook.save(filename=result_path + self.ext_xlsx)

    def make_dict_to_excel(self, path, merge_dict):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="index")
        sheet.cell(row=row, column=2, value='TTTT_Barcode')
        sheet.cell(row=row, column=3, value='Original sequence')
        sheet.cell(row=row, column=4, value='Edited sequence')
        sheet.cell(row=row, column=5, value='TTTT_Barcode_cnt')

        for barcd_key, val_dict in merge_dict.items():
            row += 1
            sheet.cell(row=row, column=1, value=str(row - 1))
            sheet.cell(row=row, column=2, value=barcd_key)
            sheet.cell(row=row, column=3, value=val_dict['Original sequence'])
            sheet.cell(row=row, column=4, value=val_dict['Edited sequence'])
            sheet.cell(row=row, column=5, value=val_dict['TTTT_Barcode_cnt'])

        workbook.save(filename=path + self.ext_xlsx)

    """
    get file lists in target dir by target ext
    :param
        path : target dir + "*." + target ext
    :return
        ['target dir/file_name.target ext', 'target dir/file_name.target ext' ...]
    """
    def get_files_from_dir(self, path):
        return glob.glob(path)

    def make_4seq_dict_to_excel(self, path, merge_dict):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="index")
        sheet.cell(row=row, column=2, value='TTTT_Barcode')
        sheet.cell(row=row, column=3, value='TTTT_Barcode_cnt')
        sheet.cell(row=row, column=4, value='Target sequences without edit')
        sheet.cell(row=row, column=5, value='Target sequences with edit (complete)')
        sheet.cell(row=row, column=6, value='Position 1 only')
        sheet.cell(row=row, column=7, value='Position 2 only')
        sheet.cell(row=row, column=8, value='ETC')

        for barcd_key, val_dict in merge_dict.items():
            row += 1
            tttt_brcd = val_dict['TTTT_Barcode_cnt']
            wout_edit_seq = val_dict['Target_sequences_without_edit']
            edit_seq = val_dict['Target_sequences_with_edit_complete']
            pos_1_seq = val_dict['Position_1_only']
            pos_2_seq = val_dict['Position_2_only']
            sheet.cell(row=row, column=1, value=str(row - 1))
            sheet.cell(row=row, column=2, value=barcd_key)
            sheet.cell(row=row, column=3, value=tttt_brcd)
            sheet.cell(row=row, column=4, value=wout_edit_seq)
            sheet.cell(row=row, column=5, value=edit_seq)
            sheet.cell(row=row, column=6, value=pos_1_seq)
            sheet.cell(row=row, column=7, value=pos_2_seq)
            sheet.cell(row=row, column=8, value=str(tttt_brcd - (wout_edit_seq + edit_seq + pos_1_seq + pos_2_seq)))

        workbook.save(filename=path + self.ext_xlsx)

    def make_4seq_dict_to_excel_(self, path, merge_dict):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="index")
        sheet.cell(row=row, column=2, value='File_Name')
        sheet.cell(row=row, column=3, value='Barcode_cnt')
        sheet.cell(row=row, column=4, value='WT sequence_cnt')
        sheet.cell(row=row, column=5, value='Prime edited_cnt')
        sheet.cell(row=row, column=6, value='WT sequence_reversed_comp_cnt')
        sheet.cell(row=row, column=7, value='Prime edited_reversed_comp_cnt')

        for barcd_key, val_dict in merge_dict.items():
            row += 1
            tttt_brcd = val_dict['TTTT_Barcode_cnt']
            wout_edit_seq = val_dict['Target_sequences_without_edit']
            edit_seq = val_dict['Target_sequences_with_edit_complete']
            pos_1_seq = val_dict['Position_1_only']
            pos_2_seq = val_dict['Position_2_only']
            sheet.cell(row=row, column=1, value=str(row - 1))
            sheet.cell(row=row, column=2, value=barcd_key)
            sheet.cell(row=row, column=3, value=tttt_brcd)
            sheet.cell(row=row, column=4, value=wout_edit_seq)
            sheet.cell(row=row, column=5, value=edit_seq)
            sheet.cell(row=row, column=6, value=pos_1_seq)
            sheet.cell(row=row, column=7, value=pos_2_seq)

        workbook.save(filename=path + self.ext_xlsx)

    def read_tsv_ignore_N_line(self, path, n_line=1, deli_str="\t"):
        result_list = []
        with open(path, "r") as f:
            for ignr_line in range(n_line):
                header = f.readline()
                print(header)
            while True:
                tmp_line = f.readline().replace("\n", "")
                if tmp_line == '':
                    break

                result_list.append(tmp_line.split(deli_str))
        return result_list

    """
    :param
        init_split_file = {'big_file_path': WORK_DIR + FASTQ + TRGT_FASTQ_NM + TRGT_FASTQ_EXT
                            , 'num_row': 4000000
                            , 'splited_files_dir': WORK_DIR + FASTQ + TRGT_FASTQ_NM + "/"
                            , 'output_file_nm': TRGT_FASTQ_NM
                            , 'output_file_ext': TRGT_FASTQ_EXT
                       }
    """
    def split_big_file_by_row(self, init):
        print('options', str(init))
        big_file_path = init['big_file_path']
        num_row = init['num_row']
        splited_files_dir = init['splited_files_dir']
        output_file_nm = init['output_file_nm']
        output_file_ext = init['output_file_ext']

        os.makedirs(splited_files_dir, exist_ok=True)

        with open(big_file_path) as fin:
            fout = open('{}/{}_{}{}'.format(splited_files_dir, output_file_nm, '0', output_file_ext), "w")
            for i, line in enumerate(fin):
                fout.write(line)
                if (i + 1) % num_row == 0:
                    fout.close()
                    fout = open('{}/{}_{}{}'.format(splited_files_dir, output_file_nm, str(i // num_row + 1), output_file_ext), "w")

            fout.close()